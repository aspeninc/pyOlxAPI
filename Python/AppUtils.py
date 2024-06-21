"""
Common utilities
"""
from __future__ import print_function

__author__    = "ASPEN Inc."
__copyright__ = "Copyright 2024, Advanced System for Power Engineering Inc."
__license__   = "All rights reserved"
__category__  = "Common"
__pyManager__ = "no"
__email__     = "support@aspeninc.com"
__status__    = "Realease"
__version__   = "2.1.8"

import logging
import traceback
#
import sys,os
import time,csv
import subprocess,threading
import tempfile,random
import string,math
from datetime import datetime

#
try: #python 3
    import winreg
    import tkinter as tk
    import tkinter.filedialog as tkf
    import tkinter.messagebox as tkm
    from tkinter import ttk
    from tkinter import simpledialog
except: #python 2
    import _winreg as winreg
    import Tkinter as tk
    import tkFileDialog as tkf
    import tkMessageBox as tkm
    import ttk

#
def iniInput(usage):
    import argparse
    PARSER_INPUTS = argparse.ArgumentParser(epilog= "")
    PARSER_INPUTS.usage = usage
    return PARSER_INPUTS
#
def parseInput(PARSER_INPUTS,demo=0,pi=0):
    PARSER_INPUTS.add_argument('-ut'  , help = ' (int) $ASPEN internal parameter$ unit test [0-ignore, 1-unit test]', default = 0,type=int,metavar='') #key in argparse.py
    if demo>0:
        PARSER_INPUTS.add_argument('-demo', help = ' (int) demo [0-ignore, 1-run demo]', default = 0,type=int,metavar='')
    if pi>0:
        PARSER_INPUTS.add_argument('-pi'  , help = ' (str) $ASPEN internal parameter$', default = [],nargs='+',metavar='') #key in argparse.py
    #
    PARSER_INPUTS.add_argument('-olxpath' , help = ' (str) Full pathname of the folder, where the ASPEN olxapi.dll is located',default = '',type=str,metavar='')
    #
    ARGVS = PARSER_INPUTS.parse_known_args()[0]
    return ARGVS
#
def encode3(s):
    """
    convert string => bytes
    for compatibility python 2/3
    """
    if type(s)==str:
        return s.encode('UTF-8')
    return s
#
def decode(s):
    """
    convert bytes => string
    """
    if type(s)==bytes:
        return s.decode('UTF-8')
    return s
#
def getASPENFile(path,sfile):
    """
    return: abs path of sfile
            None if not found
    """
    if not os.path.isfile( os.path.join(path,sfile) ) :
        path = "C:\\Program Files (x86)\\ASPEN\\1LPFv15"
    if not os.path.isfile( os.path.join(path,sfile) ) :
        path = "C:\\Program Files\\ASPEN\\1LPFv15"
    if not os.path.isfile( os.path.join(path,sfile) ) :
        path = "C:\\Program Files (x86)\\ASPEN\\1LPFv14\\OlrxAPI"
    if not os.path.isfile( os.path.join(path,sfile) ) :
        path = "C:\\Program Files\\ASPEN\\1LPFv14\\OlrxAPI"
    #
    sf = os.path.join(path,sfile)
    if os.path.isfile(sf) :
        return os.path.abspath(sf)
    else:
        return None
#
def logger2File(PY_FILE,prt=True,flog='',version=''):
    if not flog.endswith('.log'):
        flog = os.path.join(get_opath(""), os.path.split(os.path.abspath(PY_FILE))[1])
        flog = get_file_out(fo=flog , fi='' , subf='' , ad='' , ext='.log')
    flog = os.path.abspath(flog)
    #
    if prt:
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)
        #
        logging.basicConfig(level=logging.INFO,
                format='%(message)s', #%(levelname)s
                handlers=[logging.StreamHandler(sys.stdout),logging.FileHandler(flog)])
    else:
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)
        #
        logging.basicConfig(level=logging.INFO,
                format='%(message)s', #%(levelname)s
                handlers=[logging.FileHandler(flog)])
    #
    pe = os.path.split(sys.executable)[1]
    if pe in ['python.exe','pythonw.exe']:
        logging.info('Run : ' + PY_FILE + ' ' + version)
    else:
        logging.info('Run : ' + pe + ' ' +version)
    #
    logging.info('User: ' + os.getlogin())
    logging.info('Date: ' + time.asctime())
    logging.info('logFile: ' + flog)
    return flog
#
def loggerOffStdout():
    try:
        flog= logging.root.handlers[1].baseFilename
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)
        #
        logging.basicConfig(level=logging.INFO,
                            format='%(message)s', #%(levelname)s
                            handlers=[logging.FileHandler(flog)])
    except:
        pass
#
def get_String_random(k):
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=k))
#
def get_tpath():
    pf =  get_String_random(7)
    #
    tpath = get_opath('')
    tpath = os.path.join(tpath,pf)
    try:
        os.mkdir(tpath)
    except:
        pass
    return tpath
#
def get_opath(opath):
    if opath =='':
        opath= os.path.join(tempfile.gettempdir(),'1LPF')
    if not os.path.exists(opath):
        try:
            os.mkdir(opath)
        except:
            pass
    return opath
#
def get_file_out(fo,fi,subf,ad,ext):
    """
    get name file output
        fo: name given
        fi: file input (.OLR for example)
        subf: sub folder
        ad: add in the end of file
        ext: extension file output

        check if can write in folder,
        if not=> create in tempo directory
    """
    fo = correctNameFile(fo)
    fi = correctNameFile(fi)
    if fo=='':
        fo1,ext1 = os.path.splitext(fi)
    else:
        fo1,ext1 = os.path.splitext(fo)
        subf = ''
        ad = ''
    #
    if ext=='':
        ext = ext1
    #
    path,sfile = os.path.split(fo1)
    if path=='':
        path = os.path.split(fi)[0]
        if path=='':
            path,sfile = os.path.split(os.path.abspath(fo1))
    if not os.path.exists(path):
        try:
            os.mkdir(path)
        except:
            pass
    # test folder
    if subf!='':
        path = os.path.join(path,subf)
        if not os.path.exists(path):
            try:
                os.mkdir(path)
            except:
                pass
    #
    try:
        srandom = get_String_random(15)
        sf = os.path.join(path,srandom)
        ffile = open(sf, 'w+')
        ffile.close()
        deleteFile(sf)
    except:# create in tempo directory
        path = get_opath('')
        if subf!='':
            path = os.path.join(path,subf)
            if not os.path.isdir(path):
                os.mkdir(path)
    # test file
    for k in range(1000):
        if k==999:
            raise Exception('Error in AppUtils.get_file_out(...)')
        if k>0:
            ad1 = ad + '_'+str(k)
        else:
            ad1 = ad
        #
        fo = os.path.join(path,sfile + ad1 + ext)
        #
        deleteFile(fo)
        if not os.path.isfile(fo):
            return os.path.abspath(fo)
#
def getFullPath(fi,pathF):
    path = os.path.dirname(fi)
    if path=="":
        fo1 = os.path.join(pathF,fi)
    else:
        fo1 = os.path.abspath(fi)
    return fo1
#
def saveString2File(nameFile,sres):
    #
    text_file = open(nameFile, mode='w+')
    text_file.write(sres)
    text_file.close()
    return nameFile
#
def insertString2File(nameFile,sres,index):
    with open(nameFile, 'r') as f:
        contents = f.readlines()
        contents.insert(index, sres)

    with open(nameFile, 'w') as f:
        contents = "".join(contents)
        f.write(contents)
#
def saveArString2File(nameFile,ares):
    #
    text_file = open(nameFile, mode='w+')
    for s1 in ares:
        text_file.write(s1+"\n")
    text_file.close()
    return nameFile
#
def save2CSV(nameFile,ares,delim):
    #
    with open(nameFile, mode='w') as f:
        ew = csv.writer(f, delimiter=delim, quotechar='"',lineterminator="\n")
        for a1 in ares:
            ew.writerow(a1)
        f.close()
    return nameFile
#
def add2File(nameFile,sres):
    """
    append String to a file
    """
    try:
        text_file = open(nameFile, mode= 'a')
        text_file.write(sres)
        text_file.close()
    except OSError:
        raise Exception(OSError.strerror)
#
def add2CSV(nameFile,ares,delim):
    """
    append array String to a file CSV
    """
    try:
        with open(nameFile, mode='a') as f:
            ew = csv.writer(f, delimiter=delim, quotechar='"',lineterminator="\n", quoting=csv.QUOTE_MINIMAL)
            for a1 in ares:
                ew.writerow(a1)
    except OSError:
        raise Exception(OSError.strerror)
#
def read_File_text_0(sfile):
    """
    read file text to Array String/ String
    """
    ar = []
    sres = ''
    try:
        try:# python 3
            ins = open( sfile, mode="r", encoding='UTF-8')
        except:# python 2
            ins = open( sfile, mode="r")
        #
        for line in ins:
            ar.append(line.replace("\n",""))
            sres+=line
        ins.close()
    except OSError:
        raise Exception(OSError.strerror)
    #
    return ar,sres

#
def read_File_text (sfile):
    """
    read file text to Array String
    """
    ar = []
    try:
        try:# python 3
            ins = open( sfile, mode="r", encoding='UTF-8')
        except:# python 2
            ins = open( sfile, mode="r")
        #
        for line in ins:
            ar.append(line.replace("\n",""))
        ins.close()
    except OSError:
        raise Exception(OSError.strerror)
    #
    return ar
#
def read_File_text_1 (sfile):
    """
    read file text to str
    """
    sres = ""
    try:
        try:# python 3
            ins = open(sfile, mode='r', encoding='UTF-8')
        except: #python 2
            ins = open(sfile, mode='r')
        #
        sres = ins.read()
        ins.close()
    except OSError:
        raise Exception(OSError.strerror)
    #
    return sres
#
def read_File_text_2 (sfile,linefrom):
    """
    read file text to String
        start from line (linefrom)
    """
    sres = ''
    sr1 = read_File_text(sfile)
    for i in range(linefrom,len(sr1)):
        sres += sr1[i] +"\n"
    return sres
#
def read_File_text_3 (sfile,lineTo):
    """
    read file text to Array String
    up to LineTo
    """
    ar = []
    try:
        try:# python 3
            ins = open( sfile, mode="r", encoding='UTF-8')
        except:# python 2
            ins = open( sfile, mode="r")
        #
        i = 0
        for line in ins:
            i+=1
            ar.append(line.replace("\n",""))
            if i>lineTo:
                break
        ins.close()
    except OSError:
        raise Exception(OSError.strerror)
    #
    return ar
#
def read_File_csv(fileName, delim):
    """
    read file csv to array
    """
    res = []
    with open(fileName, mode= 'r') as f:
        reader = csv.reader(f, delimiter=delim)
        for row in reader:
            res.append(row)
    return res
#
def deleteFile(sfile):
    try:
        if os.path.isfile(sfile):
            os.remove(sfile)
    except:
        pass
#
def compare2FilesText(file1,file2):
    """
    compare 2 file Text (for unit test)
    returns: arrays of number of line different
    """
    ar1 = read_File_text(file1)
    ar2 = read_File_text(file2)
    dif = []
    for i in range(max(len(ar1),len(ar2))):
        a1 = ""
        a2 = ""
        try:
            a1 = ar1[i]
        except:
            pass
        try:
            a2 = ar2[i]
        except:
            pass
        #
        a1 = a1.strip()
        a2 = a2.strip()
        if a1!=a2:
            dif.append(i+1)
    return dif

#
def unit_test_compare(PATH_FILE,PY_FILE,sres):
    """
    tool for unit test
    - save result to file.txt
    - compare result with REF
    """
    pyFile = os.path.join(PATH_FILE,PY_FILE)
    fileRes = get_file_out(fo='' , fi=pyFile, subf='' , ad='_ut' , ext='.txt')
    s1 = "UNIT TEST: " + PY_FILE +"\n"+sres
    saveString2File(fileRes,s1)
    #
    # compare
    fileREF,_ = os.path.splitext(pyFile)
    fileREF+= '_ut_REF.txt'
    #
    dif = compare2FilesText(fileREF,fileRes) # return array of line number with difference

    if len(dif)==0:
        deleteFile(fileRes)
        s = "\nPASS unit test: "+ PY_FILE+ "("+ os.path.basename(fileRes)+ "=="+os.path.basename(fileREF)+")"
        print(s)
        return True
    else:
        s = "\nPROBLEM unit test: " + PY_FILE+ "("+ os.path.basename(fileRes)+ "!="+os.path.basename(fileREF)+")"
        s+="\n\tdifferences in lines: "+str(dif)
        print(s)
        return False
#
def checkFileSelected(sfile,mes):
    if sfile=='' or os.path.isdir(sfile):
        raise Exception(mes+ ' : No file selected')
    #
    if not os.path.isfile(sfile):
        raise Exception ('FileNotFoundError: '+sfile)


def correctNameFile(fi):
    res = str(fi).strip()
    if (res.startswith('"') and res.endswith('"')) or (res.startswith("'") and res.endswith("'")):
        res = res[1:-1]
    return res
#
def checkFile(fi,exta,desc):
    fi = correctNameFile(fi)
    if fi=='':
        return False,'\n'+ desc+' : None selected'
    if not os.path.isfile(fi):
        return False,'\n'+ desc+' : File not found\n"%s"'%fi
    #
    ext = (os.path.splitext(fi)[1]).upper()
    if ext not in exta:
        return False,'\n'+desc+" : Unsupported file format " +str(ext) +'\n"%s"'%os.path.abspath(fi)
    #
    return True,''
#
def checkInputOLR(folr,pyFile,PARSER_INPUTS):
    a,s = checkFile(folr,[".OLR"],'OLR file path')
    if not a:
        if pyFile.upper().endswith('.PY'):
            return FinishCheck(pyFile,s,PARSER_INPUTS)
        else:
            raise Exception(s)
    return True
#
def checkInputFile(fi,pyFile,PARSER_INPUTS,ext,mes):
    a,s = checkFile(fi,[ext],mes)
    if not a:
        if pyFile.upper().endswith('.PY'):
            return FinishCheck(pyFile,s,PARSER_INPUTS)
        else:
            raise Exception(s)
    return True
#
def checkInputPK(pk,sobj,pyFile,PARSER_INPUTS):
    if not pk:
        if pyFile.upper().endswith('.PY'):
            return FinishCheck(pyFile,"\nNo "+sobj+" is selected. Unable to continue.",PARSER_INPUTS)
        else:
            raise Exception("\nNo "+sobj+" is selected. Unable to continue.")
    #
    return True
#
def convert_LengthUnit(len_unit_to,len_unit_from):
    """
    convert length unit from [len_unit_from] => [len_unit_to]
    """
    if len_unit_to =="" or len_unit_to == len_unit_from:
        return len_unit_from,1
    #
    dictC = {"ft_km":0.0003048, "kt_km": 1.852  , "mi_km":1.609344, "m_km":0.001, \
             "km_ft":3280.839895, "km_kt": 0.539956803, "km_mi":0.6213711922, "km_m":1000 , "km_km":1}
    #
    s1 = len_unit_from +"_km"
    s2 = "km_"+len_unit_to
    s1 = s1.replace(' ','')
    s2 = s2.replace(' ','')
    val = dictC[s1] * dictC[s2]
    #
    return len_unit_to, val
#
def convert_LengthUnit_1(len_unit_to,len_unit_from):
    return convert_LengthUnit(len_unit_to,len_unit_from)[1]
#
def chekPythonVersion(app):
    if sys.version_info[0]<3 or (sys.version_info[0]== 3 and sys.version_info[1]<4):# python >=3.4
        sTitle = "Error python version (found: " + sys.version[:6]+")"
        sMain  = "Python version >=3.4 is required \nfor running "+app
        gui_error(sTitle,sMain)
        raise Exception("\n"+sTitle + "\n"+sMain)
#
def setIco(root,path,fileIco):
    fileIco = getASPENFile(path,fileIco)
    if fileIco!=None:
        root.wm_iconbitmap(fileIco)
    return fileIco
#
def setIco_1Liner(root, png='1LinerPython.png'):
    f1 = r'C:\Program Files (x86)\ASPEN\1LPFv15\OlxAPI\Python\%s'%png
    if not os.path.isfile(f1):
        f1 = os.path.dirname(sys.executable)+r'\OlxAPI\Python\%s'%png
        if not os.path.isfile(f1):
            f1 = os.path.dirname(sys.executable)+r'\%s'%png
    try:
        root.iconphoto(True, tk.PhotoImage(file = f1))
        return
    except:
        pass
    icon = setIco(root,r'C:\Program Files (x86)\ASPEN\1LPFv15\OlxAPI\Python',"1LINER.ico")
    if icon==None:
        icon = setIco(root,r'C:\Program Files (x86)\ASPEN\1LPFv15\OlxAPI\Python',"1LINER.ICO")
    if icon==None:
        icon = setIco(root,os.path.dirname(sys.executable),"1LINER.ico")
    if icon==None:
        icon = setIco(root,os.path.dirname(sys.executable),"1LINER.ICO")

#
def gui_info(sTitle,sMain):
    if is_unittest():
        return
    print(sMain)
    root = tk.Tk()
    setIco_1Liner(root)
    #
    root.attributes('-topmost', True) # put in front of window
    root.withdraw()
    tkm.showinfo(sTitle,sMain)
    root.destroy()

def is_unittest():
    cwd = os.getcwd()
    if os.path.split(cwd)[1]=='ut' and sys.executable.endswith('python.exe') \
        and os.path.isfile(cwd+'\\ut.py') and os.path.isfile(cwd+'\\ut_OlxObj.py')\
        and os.path.isdir(cwd+'\\REF') and os.path.isdir(cwd+'\\REF2'):
        return True
    return False
#
def gui_askquestion(sTitle,sMain):
    if is_unittest():
        return 'YES'
    root = tk.Tk()
    setIco_1Liner(root)
    #
    print(sMain)
    root.attributes('-topmost', True) # put in front of window
    root.withdraw()
    choice = tkm.askquestion(sTitle, sMain)
    root.destroy()
    print(choice)
    #
    return choice.upper()
#
def gui_error(sTitle,sMain):
    if is_unittest():
        return
    #
    print(sMain)
    root = tk.Tk()
    setIco_1Liner(root)
    #
    root.attributes('-topmost', True) # put in front of window
    root.withdraw()
    tkm.showerror(sTitle,sMain)
    root.destroy()
#
def gui_saveAsExcelCSV():
    """
    gui ask to save as Excel
    """
    files = [('Excel Workbook', '*.xlsx'),('CSV(Comma delimited)', '*.csv'),('All Files', '*.*')]
    try:
        v1 = tkf.asksaveasfile(defaultextension=".xlsx",filetypes=files)
        if not (v1 is None):
            return str(v1.name)
        return ""
    except:
        tkm.showinfo("Permission denied", "Select file to save calculation results")
        return gui_saveAsExcelCSV()
#
def gui_saveAsExcel(title='Select Excel file to save calculation results'):
    """
    gui ask to save as Excel
    """
    root = tk.Tk()
    setIco_1Liner(root)
    #
    root.attributes('-topmost', True) # put in front of window
    root.withdraw()
    #
    files = [('Excel Workbook', '*.xlsx'),('All Files', '*.*')]
    try:
        v1 = tkf.asksaveasfile(title=title,defaultextension=".xlsx",filetypes=files)
        if not (v1 is None):
            root.destroy()
            return str(v1.name)
        root.destroy()
        return ""
    except:
        tkm.showinfo("Permission denied", title)
        return gui_saveAsExcel()
#
def deleteSpace1(s1):
    """
    delete space in String
        sample:  "  abc   x  y z   " => "abc x y z"
    """
    sa = (s1.strip()).rsplit()
    sres = ""
    for i in range(len(sa)):
        if i>0:
            sres +=' '
        sres += sa[i]
    return sres
#
def deleteLineBlank(s1):
    a1 = s1.split('\n')
    sres = ''
    for si in a1:
        if si.rstrip():
            sres+=si+'\n'
    return sres
#
def getValIfDictInString(s0,dictC):
    """
    example get connection code in FaultDescription
        Bus Fault on: "6 NEVADA 132. kV 3LG" => "3LG"
    """
    for key in dictC:
        if key in s0:
            return dictC[key]
    return ''
#
def getFltConn_a(sa):
    """
    array of getFltConn
    """
    res = []
    for s1 in sa:
        res.append(getFltConn(s1))
    return res
#
def array2String(ar,delim):
    """
    convert array value to String with delim
    """
    res = ""
    for i in range(len(ar)):
        if i>0:
            res +=delim
        res += str(ar[i])
    return res
#
def remove_button(tk1):
    """
     remove python as Window
    """
    from ctypes import wintypes,windll
    GWL_STYLE = -16
    WS_CHILD = 0x40000000
    WS_SYSMENU = 0x00080000

    SWP_FRAMECHANGED = 0x0020
    SWP_NOACTIVATE = 0x0010
    SWP_NOMOVE = 0x0002
    SWP_NOSIZE = 0x0001
    # write short names for functions and specify argument and return types
    tk1.GetWindowLong = windll.user32.GetWindowLongW
    tk1.GetWindowLong.restype = wintypes.ULONG
    tk1.GetWindowLong.argtpes = (wintypes.HWND, wintypes.INT)

    tk1.SetWindowLong = windll.user32.SetWindowLongW
    tk1.SetWindowLong.restype = wintypes.ULONG
    tk1.SetWindowLong.argtpes = (wintypes.HWND, wintypes.INT, wintypes.ULONG)

    tk1.SetWindowPos = windll.user32.SetWindowPos
    hwnd = int(tk1.wm_frame(), 16)
    style = tk1.GetWindowLong(hwnd, GWL_STYLE) # get existing style
    style = style & ~WS_SYSMENU
    res = tk1.SetWindowLong(hwnd, GWL_STYLE, style)
    res = tk1.SetWindowPos(hwnd, 0, 0,0,0,0, SWP_FRAMECHANGED | SWP_NOACTIVATE | SWP_NOMOVE | SWP_NOSIZE)

def hide_minimize_maximize(tk1):
    #   shortcuts to the WinAPI functionality
    from ctypes import windll
    set_window_pos = windll.user32.SetWindowPos
    set_window_long = windll.user32.SetWindowLongA
    get_window_long = windll.user32.GetWindowLongA
    get_parent = windll.user32.GetParent

    #   some of the WinAPI flags
    GWL_STYLE = -16

    WS_MINIMIZEBOX = 131072
    WS_MAXIMIZEBOX = 65536

    SWP_NOZORDER = 4
    SWP_NOMOVE = 2
    SWP_NOSIZE = 1
    SWP_FRAMECHANGED = 32


    hwnd = get_parent(tk1.winfo_id())
    #   getting the old style
    old_style = get_window_long(hwnd, GWL_STYLE)
    #   building the new style (old style AND NOT Maximize AND NOT Minimize)
    new_style = old_style & ~ WS_MAXIMIZEBOX & ~ WS_MINIMIZEBOX
    #   setting new style
    set_window_long(hwnd, GWL_STYLE, new_style)
    #   updating non-client area
    set_window_pos(hwnd, 0, 0, 0, 0, 0, SWP_NOMOVE | SWP_NOSIZE | SWP_NOZORDER | SWP_FRAMECHANGED)

#
def runSubprocess(args,cwd=''):
    """
    run a subprocess file with arg
    """
    if cwd:
        proc = subprocess.Popen(args, stdin=subprocess.PIPE, stdout=subprocess.PIPE,stderr=subprocess.PIPE,cwd=cwd)
    else:
        proc = subprocess.Popen(args, stdin=subprocess.PIPE, stdout=subprocess.PIPE,stderr=subprocess.PIPE)
    outb, errb = proc.communicate()
    out = outb.decode('UTF-8','ignore')
    err = errb.decode('UTF-8','ignore')
    #
    return out,err,proc.returncode
#
def runSubprocess_withError(args,cwd=''):
    """
    run a subprocess file with arg
      raise Exception if error
    """
    out,err,returncode = runSubprocess(args,cwd)
    if returncode>0:
        raise Exception(err)
    return out,err,returncode
#
def getRootError(err):
    a = err.split('\n')
    sr = ''
    listError = ['Exception ','raise ','AttributeError','ImportError','IndexError','KeyError','TypeError','ValueError']

    for i in range(len(a)):
        si = (a[len(a)-i-1]).strip()
        for le in listError:
            if si.startswith(le):
                sr = si +'\n' +sr
                sr = sr.replace('Exception:','')
                sr = sr.replace('Exception(err)','')
                sr = sr.replace('raise','')
                return sr.strip()
        sr = si +'\n' +sr
    return sr
#
def FinishException(pyFile,err):
    if pyFile.upper().endswith('.PY'):
        try:
            flog= logging.root.handlers[1].baseFilename
        except:
            flog = logger2File(pyFile,prt=False)
        logging.info('\n'+traceback.format_exc())
        print('\nlogFile',flog)
        print('\nERROR: '+str(err))
        #
        logging.shutdown()
    else:
        raise Exception(err)
#
def FinishCheck(pyFile,err,PARSER_INPUTS=None):
    loggerOffStdout()
    logging.info('\n'+err)
    if PARSER_INPUTS!=None:
        PARSER_INPUTS.print_help()
    print('\nERROR:'+err)
    #
    logging.shutdown()
    return False
#
def runSubprocess_noWait(args):
    """
    run subprocess without waiting
    """
    return subprocess.call(args)
#
def runSubprocess_getHelp(pyFile, pythonPath='',cwd='',timeout=1):
    """
    get help of python (-h) by subprocess
    """
    pathe = pythonPath if pythonPath else os.path.dirname(sys.executable)
    args = [os.path.join(pathe,'pythonw.exe'),pyFile,'-h']
    result = subprocess.run(args, cwd=cwd,timeout=timeout,capture_output=True)
    out = result.stdout.decode('UTF-8','ignore')
    err = result.stderr.decode('UTF-8','ignore')
    #
    return out,err,err!=''
#
def runSubprocess_getHelp2(pyFile, pythonPath='',cwd='',timeout=1):
    """
    get help of python (-h) by subprocess
    """
    pathe = pythonPath if pythonPath else os.path.dirname(sys.executable)
    args = [os.path.join(pathe,'pythonw.exe'),pyFile,'-h']
    result = subprocess.run(args, cwd=cwd,timeout=timeout,capture_output=True)
    out = result.stdout.decode('UTF-8','ignore')
    err = result.stderr.decode('UTF-8','ignore')
    #
    return out,err,err!=''
#
def runSubprocess_1(pyFile,args, pythonPath=''):
    """
    run python by subprocess with args
    """
    pathe = pythonPath if pythonPath else os.path.dirname(sys.executable)
    ARGSR = [os.path.join(pathe,'python.exe'),pyFile]
    ARGSR.extend(args)
    #
    return runSubprocess(ARGSR)
#
def runSubprocess_1w(pyFile,args, pythonPath='',cwd=''):
    """
    run pythonw by subprocess with args
    """
    pathe = pythonPath if pythonPath else os.path.dirname(sys.executable)
    ARGSR = [os.path.join(pathe,'pythonw.exe'),pyFile]
    ARGSR.extend(args)
    #
    return runSubprocess(ARGSR,cwd)
#
def runSubprocess_1w_withError(pyFile,args, pythonPath='',cwd=''):
    """
    run a subprocess file with arg pythonw.exe
      raise Exception if error
    """
    out,err,returncode = runSubprocess_1w(pyFile,args, pythonPath=pythonPath,cwd=cwd)
    if returncode>0:
        raise Exception(err)

#
def launch_OneLiner(fo,olxpath=''):
    sf = getASPENFile(olxpath,"oneline.exe")
    if sf!=None:
        path1= os.path.dirname(sf)
        os.chdir(path1)
        #
        s1 = 'start oneline.exe "{0}"'.format(fo)
        os.system(s1)
#
def launch_OneLiner_ask(fo,sTitle,sMain):
    sMain2 = "Do you want to open it in OneLiner?"
    sMain1 = sMain + fo
    choice = gui_askquestion(sTitle,sMain1+"\n\n" +sMain2)
    if choice=='yes':
        launch_OneLiner(fo)
#
def launch_excel(sfile):
    """
    Openning excel file (sfile)
    check correct path, .xls, .xlsx, .csv
    """
    ck = checkFileType(fi=sfile,ext=['.CSV','.XLSX','.XLSM','.XLS'],err=True,sTitle0="Openning excel")
    #
    os.system('start "excel.exe" "{0}"'.format(sfile))
#
def launch_excel_ask(sfile,sTitle,sMain):
    sMain2 = "Do you want to open it in excel?"
    sMain1 = sMain + sfile
    choice = gui_askquestion(sTitle,sMain1+"\n\n" +sMain2)
    if choice=='yes':
        launch_excel(sfile)
#
def launch_notepad(sfile):
    # os.system('start notepad.exe "{0}"'.format(sfile))
    subprocess.Popen(["notepad.exe", sfile])

#
def launch_notepad_ask(sfile,sTitle,sMain):
    sMain2 = "Do you want to open it in notepad?"
    sMain1 = sMain + sfile
    choice = gui_askquestion(sTitle,sMain1+"\n\n" +sMain2)
    if choice=='yes':
        launch_notepad(sfile)
#
def launch_notepadpp(sfile):
    os.system('start "notepad.exe" "{0}"'.format(sfile))
#
def launch_notepadpp_ask(sfile,sTitle,sMain):
    sMain2 = "Do you want to open it in notepadpp?"
    sMain1 = sMain + sfile
    choice = gui_askquestion(sTitle,sMain1+"\n\n" +sMain2)
    if choice=='yes':
        launch_notepadpp(sfile)
#
def checkFileType(fi,ext,err,sTitle0):
    """
    check file is with extension [ext]
       if err= True => GUI error if file not found or file don't have extension in ex
       if not => return -10: file not found
                 return -1 : file don't have extension in ex
                 return i: index extension file
    """
    if not os.path.isfile(fi):
        if err:
            sMain = "No such file or directory:\n"+fi
            raise Exception(sTitle0+sMain)
        else:
            return -10
    #
    ext1 = (os.path.splitext(fi)[1]).upper()
    for i in range(len(ext)):
        if ext1 ==ext[i]:
            return i
    #
    if err:
        sMain = "\nError format file (" +str(ext)+ ") : " +fi
        raise Exception(sTitle0+sMain)
    else:
        return -1
#
def getValNumber(row):
    ar = []
    for v1 in row:
        if type(v1) not in {int ,float}:
            try:
                v1 = int(v1)
            except:
                try:
                    v1 = float(v1)
                except:
                    pass
        ar.append(v1)
    return ar
#
def exitApplication(app):
    """
        exitApplication("python.exe")
        exitApplication("EXCEL.exe")
    """
    try:
        os.system("taskkill /im {0}".format(app))
        os.system('taskkill /f /im {0}'.format(app))
    except:
        pass
#
class TraceThread(threading.Thread):
    def __init__(self, *args, **keywords):
        threading.Thread.__init__(self, *args, **keywords)
        self.killed = False
    def start(self):
        self._run = self.run
        self.run = self.settrace_and_run
        threading.Thread.start(self)
    def settrace_and_run(self):
        sys.settrace(self.globaltrace)
        self._run()
    def globaltrace(self, frame, event, arg):
        return self.localtrace if event == 'call' else None
    def localtrace(self, frame, event, arg):
        if self.killed and event == 'line':
            raise SystemExit()
        return self.localtrace
#
class ToolCSVExcell:
    """
    tool read excel csv using openpyxl
    """
    def __init__(self):
        #
        self.guiErr = True
    #
    def setguiErr(self,guiErr):
        self.guiErr = guiErr

    #
    def readFile(self,fi,delim):
        fi = correctNameFile(fi)
        ck = checkFileType(fi=fi,ext=[".CSV",".XLSX",".XLSM"],err=True,sTitle0="Openning Excel/CSV file")
        #
        if ck==0:
            self.readFile_csv(fi,delim)
        else:
            self.readFile_EXCEL(fi)
    #
    def readFile_csv(self,fi,delim):
        self.isExcel =  False
        fi = correctNameFile(fi)
        #read file CSV
        checkFileType(fi=fi,ext=[".CSV"],err=True,sTitle0="Openning CSV file")
        #
        self.fi = fi
        self.ws = read_File_csv(fileName=fi, delim=delim)
        self.flag = 0
        self.currentSheet = None

    # read file EXCEL
    def readFile_EXCEL(self,fi):
        from openpyxl import load_workbook
        fi = correctNameFile(fi)
        self.isExcel =  True
        checkFileType(fi=fi,ext=['.XLSX','.XLS','.XLSM'],err=True,sTitle0="Openning Excel file")
        #
        self.fi = fi
        #
        self.wb = load_workbook(fi)
        self.ws = self.wb.active
        self.currentSheet = self.ws.title
        self.flag = 1
        self.allSheet = self.wb.sheetnames[:]
        return

    def close(self):
        try:
            self.wb.close()
        except:
            pass
    #
    def getVal(self,row,column):
        if self.flag == 0: # CSV file
            try:
                return self.ws[row-1][column-1]
            except:
                return None
        #
        if self.flag == 1: # openpyxl
            return self.ws.cell(row, column).value
    #
    def selectSheet(self,nameSheet):
        if self.currentSheet == nameSheet:
            return
        #
        if (nameSheet not in self.allSheet) and (self.flag ==1):
            sMain1 = "File : " + self.fi
            sMain2 = "Sheet not found : " + nameSheet
            raise Exception(sMain1+"\n"+ sMain2)
        #
        if self.flag==1:
            self.ws = self.wb[nameSheet]
            self.currentSheet = nameSheet
    #
    def getValRow(self,row,columnFr):
        """
        get value of row from column
            break if blank is found
        """
        res = []
        i   = columnFr
        #
        while True:
            v1 = self.getVal(row,i)
            if v1==None or v1=='':
                break
            #
            res.append(v1)
            i+=1
        #
        if len(res)==0:
            res=['']
        return res
    #
    def getValRowf(self,row,columnFr,leng):
        """
        get value of row from column
            break by leng
        """
        res = []
        for i in range(leng):
            v1 = self.getVal(row,columnFr+i)
            if v1==None:
                res.append('')
            else:
                res.append(v1)
        return res
    #
    def save2Excel(self,nameFileExcel,ares,nameSheet):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        try:
            ws.title = nameSheet[0]
        except:
            pass
        #
        for a1 in ares[0]:
            ws.append(getValNumber(a1))
        #
        for i in range(1,len(ares)):
            try:
                ws = wb.create_sheet(nameSheet[i])
            except:
                ws = wb.create_sheet()
            #
            for a1 in ares[i]:
                ws.append(getValNumber(a1))
        #
        wb.save(filename=nameFileExcel)
        #
        wb.close()
#
class WIN_REGISTRY:
    def __init__(self,path,keyUser,nmax):
        #
        self.nmax = nmax
        #
        if keyUser == "LOCAL_MACHINE":
            self.Registry = winreg.ConnectRegistry(None, winreg.HKEY_LOCAL_MACHINE)
        else:#"CURRENT_USER":
            self.Registry = winreg.ConnectRegistry(None, winreg.HKEY_CURRENT_USER)
        #
        try:
            self.RawKey  = winreg.OpenKey(self.Registry, path)
        except:
            self.createKey(path)
            self.RawKey  = winreg.OpenKey(self.Registry, path)
        #
        self.reg_key = winreg.OpenKey(self.Registry,path,0, winreg.KEY_SET_VALUE)
    #
    def deleteInvalidFile(self): #delete invalid file in registry
        va,i = [],0
        while True:
            try:
                v1 = winreg.EnumValue(self.RawKey, i)
                if not os.path.isfile(v1[1]):
                    va.append(v1)
            except:
                break
            i+=1
        for v1 in va:
            winreg.DeleteValue(self.reg_key, v1[0])
    #
    def createKey(self,path):
        patha = str(path).split("\\")
        path1 = patha[0]
        for i in range(1,len(patha)):
            p2 = path1 +"\\"+ patha[i]
            try:
                access_key = winreg.OpenKey(self.Registry,p2)
            except:
                access_key = winreg.OpenKey(self.Registry,path1)
                winreg.CreateKey(access_key,patha[i])
            #
            path1 = p2
    #
    def getAllNameValue(self):
        i = 0
        name,vala = [],[]
        while True:
            try:
                a = winreg.EnumValue(self.RawKey, i)
                name.append(a[0])
                val1 = a[1]
                try:
                    val1 = val1.replace('\\','/')
                    val1 = val1.replace('//','/')
                    val1 = val1.replace('//','/')
                except:
                    pass
                vala.append(val1)
            except:
                break
            i+=1
        return name,vala
    #
    def getAllValue(self):
        return self.getAllNameValue()[1]
    #
    def getValue0(self):
        try:
            v,a = self.getAllNameValue()
            for a1 in a:
                if a1:
                    return a1
            return ""
        except:
            return ""
    #
    def appendValue(self,val):
        name,vala = self.getAllNameValue()
        if len(vala)>0 and val==vala[0]:
            return False
        #
        for n1 in name:
            winreg.DeleteValue(self.reg_key, n1)
        #
        val1 = val.replace('\\','/')
        val1 = val1.replace('//','/')
        val1 = val1.replace('//','/')
        r1 = [val1]
        for ri in vala:
            if len(r1)>=self.nmax:
                break
            if ri!=val1 :
                r1.append(ri)
        #
        for i in range(len(r1)):
            winreg.SetValueEx(self.reg_key, "File"+str(i+1), 0, winreg.REG_SZ, r1[i])
        #
        return True
    #
    def updateValue(self,sf1,val):
        try:
            winreg.SetValueEx(self.reg_key, sf1, 0, winreg.REG_SZ, val)
        except:
            pass
    #
    def getValue(self,sf1,default=''):
        name,vala = self.getAllNameValue()
        for i in range(len(name)):
            if name[i]==sf1:
                return vala[i]
        return default
    #
    def deleteValue(self,val):
        name,vala = self.getAllNameValue()
        #
        for i in range(len(name)):
            if vala[i]==val:
                winreg.DeleteValue(self.reg_key, name[i])
    #
    def deleteValueFile(self,val):
        val1 = val.replace('\\','/')
        val1 = val1.replace('//','/')
        val1 = val1.replace('//','/').upper()
        name,vala = self.getAllNameValue()
        #
        for i in range(len(name)):
            if vala[i].upper()==val1:
                winreg.DeleteValue(self.reg_key, name[i])
#
class Gui_Select(tk.Frame):
    def __init__(self, parent,w1,xOK,yOK,xCom,yCom,data,select):
        tk.Frame.__init__(self, parent)
        self.parent = parent
        self.data = data
        self.select = select
        self.initGUI(w1,xOK,yOK,xCom,yCom)
    #
    def initGUI(self,w1,xOK,yOK,xCom,yCom):
        #
        self.cb = ttk.Combobox(self.parent,width=w1, values=self.data)
        self.cb.place(x=xCom, y=yCom)
        self.cb.bind("<<ComboboxSelected>>")
        self.cb.current(0)
        #
        button_OK = tk.Button(self.parent,text =   '     OK     ',command=self.run_OK)
        button_OK.place(x=xOK, y=yOK)
    #
    def run_OK(self):
        x = self.cb.get()
        self.select.append(x)
        self.parent.destroy()
#
def compare2Val(v1,v2):
    """
    return typ, v1,v2,c
        '
        typ = 1 Number
        typ = 2 String
        typ = 0 error data
        '
        c: val compare
    """

    try:# float
        f1 = float(v1)
        f2 = float(v2)
        af1 = abs(f1)
        af2 = abs(f2)
        f12 = abs(f1-f2)
        try:
            c = f12/min(af1,af2)
        except:
            try:
                c = f12/max(af1,af2)
            except:
                c = 0
        typ = 1
        #
        return typ,f1,f2,c
    #
    except:
        try:  # String
            s1 = str(v1)
            s2 = str(v2)
            if s1==s2:
                c= ""
            else:
                c = "!"
            typ = 2
            return typ,s1,s2,c
        except:
            try:
                s1 = str(v1)
            except:
                s1 = ""
            #
            try:
                s2 = str(v2)
            except:
                s2 = ""
            #
            c = None
            typ = 0
            return typ,s1,s2,c
#
def compare2Dict(dic1,dic2,cr):
    # return dif{} at each value of dic1,dic2
    dif = dict()
    ndif = 0
    for key in dic1:
        try:
            typ,v1,v2,c = compare2Val(dic1[key],dic2[key])
            if typ==1: # Number
                if c>cr:
                    dif[key] = True
                    ndif +=1
                else:
                    dif[key] = False
            elif typ==2: # String
                if c!="":
                    dif[key] = True
                    ndif +=1
                else:
                    dif[key] = False
        except:
            pass
    return ndif,dif
#
def getMemory():
    import psutil
    pid = os.getpid()
    py = psutil.Process(pid)
    memoryUse = py.memory_info()[0]
    memoryUse = memoryUse/(2.**30)*1024
    return memoryUse
#
def printMemory():
    s = 'Memory used= %.1fMB' %(getMemory())
    logging.info(s)
#
def printImpedance(dR, dX, dKV, BaseMVA):
    dMag = math.sqrt(dR*dR + dX*dX)*dKV*dKV/BaseMVA
    try:
        dAng = math.atan(dX/dR)*180/math.pi
    except:
        if dX > 0:
            dAng = 90
        else:
            dAng = -90
    aLine = "{0:.5f}".format(dR) + "+j" + "{0:.5f}".format(dX) + "pu(" + "{0:.2f}".format(dMag) + "@" + "{0:.2f}".format(dAng) + "Ohm)"
    return aLine
#
def ask_run_demo(PY_FILE,ut,fi,ad):
    if ut<0:
        return 'yes'
    sTitle = os.path.splitext(PY_FILE)[0]+ " demo"
    sMain = ''
    if fi:
        sMain = "\nOLR file: " + fi + '\n'
    if ad:
        sMain +=  ad + '\n\n'
    sMain += "Do you want to run this demo (Y/N)?"
    print( "\n" + sTitle )
    choice = input( sMain )
    if str(choice.upper()) == 'Y':
        return 'yes'
    else:
        return 'cancel'
    #
    #choice = gui_askquestion(sTitle,sMain)
    #print(sTitle + sMain +": "+str(choice.upper()) +'\n')
    #return str(choice.upper())
#
def demo_notFound(PY_FILE,val,val0,gui_err=0):
    sMain = "\nUnsupported: "+os.path.splitext(PY_FILE)[0]+ " demo= "+ str(val)
    sMain +="\nTry demo=" +str(val0)
    print(sMain)
    if gui_err:
        gui_error(PY_FILE,sMain)
#
def isFile(path,sfile):
    ffile = os.path.join(path,sfile)
    return os.path.isfile(ffile)
#
def openFile(path,sfile):
    sfile1 = os.path.join(path,sfile)
    deleteFile(sfile1)
    try:
        ffile = open(sfile1, 'w+')
        return ffile,sfile1
    except Exception as e:
        logging.error(str(e))
        raise e
#
def closeFile(ffile):
    try:
        ffile.close()
    except Exception as e:
        logging.error(str(e))
        raise e
#
def createFile(path,sfile):
    sfile1 = os.path.join(path,sfile)
    if not os.path.isfile(sfile1):
        ffile = open(sfile1, 'w+')
        ffile.close()
#
def getStrNow():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#
def getShortNameFile(fname,ln):
    """
    D:\XXXXXXX\YYYYYYYYYYYYY\sample\SAMPLE30.OLR
    => D:\...\sample\SAMPLE30.OLR
    """
    if len(fname)<=ln:
        res = fname
        for i in range(len(fname),ln):
            res +='_'
    else:
        res = fname[0:3] + '...'
        res +=fname[len(fname)-ln+6:len(fname)]
    return res
#
def runCommand(command,PATH_FILE,verbose=1):
    t0 = time.time()
    pyFile = os.path.basename(command[1])
    pe = os.path.basename(command[0])
    if pe =='python.exe':
        returncode = runSubprocess_noWait(command)
        t1 = time.time()
        if t1-t0>5:
             print("runtime: %.1fs" %(t1- t0))
        if returncode>0:
            if verbose:
                print()
                os.system("pause")
            #gui_error("ERROR","ERROR: "+pyFile)
        else:
            createFile(PATH_FILE,"success")
            if not os.path.isfile( os.path.join(PATH_FILE,"pst.bas")):
                # print("SUCCESSFUL: "+command[1])
                if verbose:
                    print()
                    os.system("pause")
                #gui_info("SUCCESSFUL","SUCCESSFUL: "+pyFile)
    else:
        out,err, returncode = runSubprocess(command)
        #
        flog = logger2File(pyFile)
        logging.info(out)
        t1 = time.time()
        if t1-t0>5:
             logging.info("runtime: %.1fs" %(t1- t0))
        #
        if returncode!=0:
            logging.error(err)
            logging.shutdown()
            se = getRootError(err)
            if se:
                gui_error("ERROR",se+"\n\nlogFile : "+flog)
                # AppUtils.launch_notepad(flog)
        else:
            logging.shutdown()
            createFile(PATH_FILE,"success")
#
def getFileDemo(path0,sfile):
    s1 = os.getcwd()+os.path.sep+sfile
    if os.path.isfile(s1):
        return s1
    s1= os.getcwd()+os.path.sep +'sample'+os.path.sep+sfile
    if os.path.isfile(s1):
        return s1
    s1= path0+os.path.sep+sfile
    if os.path.isfile(s1):
        return s1
    s1= path0+os.path.sep+'sample'+os.path.sep+sfile
    if os.path.isfile(s1):
        return s1
    raise Exception('File demo not found: '+sfile)
#
def encodingFile(sFile):
    fnew = get_file_out(fo='' , fi=sFile , subf='' , ad='_1', ext='')
    f0 = open(sFile,"rb")# binary format read file
    f1 = open(fnew, "w")
    while True:
        line = f0.readline()
        if not line:
            break
        else:
            try:
                s = line.decode('UTF-8', 'ignore')[:-1]
                f1.write(s)
            except:
                try:
                    s = line.decode('UTF-8')[:-1]
                    f1.write(s)
                except:
                    pass
    f0.close()
    f1.close()
    return fnew
#
def isShipVersion():
    pe = os.path.basename(sys.executable)
    return (pe not in ['python.exe','pythonw.exe'])

#
def pauseFinal(verbose):
    if verbose>0 and isShipVersion():
        print()
        os.system("pause")
#
# Create a run marker file in this folder and keep it open during the
#        # entire Python execution session. OneLiner will pause execution
#        # as long as this file exists to wait for the result from Python
def markerStart(opath):
    if opath:
        deleteFile(opath+'\\success')
        deleteFile(opath+'\\cancel')
        return openFile(opath,'running')
    return None,None
#
def markerSucces(opath):
    if opath:
        createFile(opath,'success')
#
def markerStop(opath,FRUNNING,SRUNNING):
    if opath:
        if not isFile(opath,'success'):
            createFile(opath,'cancel')
        closeFile(FRUNNING)
        deleteFile(SRUNNING)
#
def saveAsIgnoreAr(fi,fo,ar):
    f0 = open(fi,"rb")# binary format read file
    f1 = open(fo, "wb")
    while True:
        line = f0.readline()
        if not line:
            break
        #
        t = True
        for a1 in ar:
            if line.count(a1)>0:
                t = False
                break
        if t:
            f1.write(line)
    f0.close()
    f1.close()

#
def calOffsetMonitors(master):
    #get_monitors:
    from ctypes import windll,Structure,WINFUNCTYPE,c_long,c_int, c_ulong, POINTER, c_double
    user = windll.user32
    class RECT(Structure):
        _fields_ = [('left', c_long), ('top', c_long), ('right', c_long),('bottom', c_long)]
    monitors = []
    CBFUNC = WINFUNCTYPE(c_int, c_ulong, c_ulong, POINTER(RECT), c_double)
    def cb(hMonitor, hdcMonitor, lprcMonitor, dwData):
        r = lprcMonitor.contents
        data = [r.left, r.top, r.right, r.bottom]#[hMonitor]
        monitors.append(data)
        return 1
    cbfunc = CBFUNC(cb)
    temp = user.EnumDisplayMonitors(0, 0, cbfunc, 0)
    #
    sw = master.winfo_screenwidth()
    sh = master.winfo_screenheight()
    currentXY = master.winfo_pointerxy()
    #
    for m1 in monitors:
        if (currentXY[0]>m1[0] and currentXY[0]<m1[2]) and (currentXY[1]>m1[1] and currentXY[1]<m1[3]):
            return m1[0], m1[1],abs(m1[0]-m1[2]),abs(m1[1]-m1[3])
    return 0,0,sw,sh
#
def setGeometry(master,szwindow):
    offX,offY,sw,sh = calOffsetMonitors(master)
    va = szwindow.split()
    if len(va)==2: # '60 20'
        x1 = float(va[0])/100
        y1 = float(va[1])/100
        x0 = 0.5-x1/2+offX/sw
        y0 = 0.5-y1/2+offY/sh
    else:#'20 60 15 60'
        x0 = float(va[0])/100+offX/sw
        x1 = float(va[1])/100
        y0 = float(va[2])/100+offY/sh
        y1 = float(va[3])/100
    master.geometry("{0}x{1}+{2}+{3}".format(int(x1*sw),int(y1*sh),int(x0*sw),int(y0*sh)))
#
def isEmbedded():
    return sys.executable.upper().endswith('ONELINE.EXE')


def explorerDir(path, sMain='', title=''):
    # open a folder on Windows Explorer
    if is_unittest():
        return
    if sMain:
        s1= sMain+"\n\nDo you want to open the folder in Windows Explorer (Y/N)?"
        choice = gui_askquestion(title,s1)
        if choice=='YES':
            explorerDir(path)
    else:
        if os.path.isfile(path):
            subprocess.run(['explorer', os.path.realpath(os.path.dirname(path))])
        elif os.path.isdir(path):
            subprocess.run(['explorer', os.path.realpath(path)])
